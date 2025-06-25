import os
import pandas as pd
import time
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook


PREFIXOS = {
    "PA1": {"CC13", "CC21", "CL2005", "CL3008", "CL3010", "CL3012", "CL3013", "CL3014",
            "CL3016", "CL3017", "CL3021", "CL3022", "CL3024", "CL3035", "CL3037",
            "CC02", "CC03", "CC09"},
    "PA2": {"CL19", "CL2002", "CL2004", "CL3002", "CL3003", "CL3004", "CL3005",
            "CL3006", "CL3009", "CL3015", "CL3036", "CL3039", "CL3042", "CL3043", "CL206"},
    "PA3": {"C11", "CL2001", "CL2003", "CL2006", "CL2018", "CL3001", "CL3011",
            "CL3018", "CL3023", "CL3025", "CL3027", "CL3030", "CL3032", "CL3033",
            "CL3024", "CL2013", "CL2014", "CL2015"},
    "PA4": {"CC04", "CC10", "CC14", "CL2007", "CL2008", "CL3007", "CL3019", "CL3020",
            "CL3026", "CL3028", "CL3029", "CL3031", "CL3038", "CL3040", "CL3041",
            "CC16", "CC17", "SL001"}
}


COLUNAS_REMOVER = [
    "CIDADE", "ESTADO", "LATITUDE", "LONGITUDE", "LIMIAR", "UO", "PONTO_REFERENCIA",
    "USUARIO FEEDBACK", "FEEDBACK VIDEO", "ULTIMO_COMENTARIO", "ATRIBUIDO",
    "CATEGORIA", "CERCA ELETRÔNICA", "INTEGRADOR", "GRUPO", "VISUALIZADO POR",
    "NIVEL", "RÓTULO"
]


def formatar_excel(path_arquivo: str):
    try:
        wb = load_workbook(path_arquivo)
        ws = wb.active
        font_bold = Font(bold=True, name="Calibri", size=11)
        
       
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = font_bold
        
        
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

        wb.save(path_arquivo)
    except Exception as e:
        print(f"⚠️ Erro ao formatar Excel {path_arquivo}: {e}")


def filtrar_dataframe_por_prefixo(df: pd.DataFrame, prefixos: set, pa: str) -> pd.DataFrame:
   
    df_filtrado = df[df["PREFIXO"].isin(prefixos)].copy()
    
   
    cols_para_remover = [col for col in COLUNAS_REMOVER if col in df_filtrado.columns]
    if cols_para_remover:
        df_filtrado.drop(columns=cols_para_remover, inplace=True)
    
    
    df_filtrado.insert(0, "PA", pa)
    return df_filtrado


def processar_arquivo(filepath: str, progress_callback=None) -> list:
  
    try:
        print(f"📂 Processando arquivo: {os.path.basename(filepath)}")
        
        
        df = pd.read_csv(filepath, encoding="latin1", sep=None, engine="python")
        print(f"📊 Dados carregados: {len(df)} registros")
        
       
        pasta_exportados = os.path.join(os.path.dirname(filepath), "exportados")
        os.makedirs(pasta_exportados, exist_ok=True)

        arquivos_gerados = []
        total_grupos = len(PREFIXOS)

       
        for i, (pa, prefixos) in enumerate(PREFIXOS.items(), start=1):
            try:
                print(f"🔄 Processando {pa}... ({i}/{total_grupos})")
                
                
                df_filtrado = filtrar_dataframe_por_prefixo(df, prefixos, pa)
                
                if df_filtrado.empty:
                    print(f"⚠️ Nenhum dado encontrado para {pa}")
                    continue
                
              
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:21]
                nome_arquivo = f"veiculos_{pa.lower()}_filtrado_{timestamp}.xlsx"
                caminho_saida = os.path.join(pasta_exportados, nome_arquivo)

                
                df_filtrado.to_excel(caminho_saida, index=False)
                formatar_excel(caminho_saida)
                arquivos_gerados.append(caminho_saida)
                
                print(f"✅ {pa}: {len(df_filtrado)} registros salvos")

                
                if progress_callback:
                    try:
                        progress_callback(i / total_grupos)
                    except Exception as callback_error:
                        print(f"⚠️ Erro no callback de progresso: {callback_error}")

                time.sleep(0.1)  
                
            except Exception as pa_error:
                print(f"❌ Erro ao processar {pa}: {pa_error}")
                continue

        
        return arquivos_gerados
        
    except Exception as e:
        return []
