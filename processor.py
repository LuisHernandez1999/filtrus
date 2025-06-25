import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import pandas as pd
import time
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import matplotlib
import matplotlib.patches as patches
from matplotlib.patches import Rectangle
# Configurar matplotlib para não usar GUI quando necessário
matplotlib.use('Agg')  # Backend sem GUI

# Configurações dos prefixos por PA
PREFIXOS = {
    "PA1": {"CC13", "CC21", "CL2005", "CL3008", "CL3010", "CL3012", "CL3013", "CL3014",
            "CL3016", "CL3017", "CL3021", "CL3022", "CL3024", "CL3035", "CL3037",
            "CC02", "CC03", "CC09"},
    "PA2": {"CL19", "CL2002", "CL2004", "CL3002", "CL3003", "CL3004", "CL3005",
            "CL3006", "CL3009", "CL3015", "CL3036", "CL3039", "CL3042", "CL3043", "CL206"},
    "PA3": {"C11", "CL2001", "CL2003", "CL2006", "CL2018", "CL3001", "CL3011",
            "CL3018", "CL3023", "CL3025", "CL3027", "CL3030", "CL3032", "CL3033",
            "CL3024", "CL2013", "CL2014", "CL2015"},
    "PA4": {"CC04", "CC10", "CC14", "CL2007", "CL2008", "CL3007", "CL3019", "CL3020",
            "CL3026", "CL3028", "CL3029", "CL3031", "CL3038", "CL3040", "CL3041",
            "CC16", "CC17", "SL001"}
}

COLUNAS_REMOVER = [
    "CIDADE", "ESTADO", "LATITUDE", "LONGITUDE", "LIMIAR", "UO", "PONTO_REFERENCIA",
    "USUARIO FEEDBACK", "FEEDBACK VIDEO", "ULTIMO_COMENTARIO", "ATRIBUIDO",
    "CATEGORIA", "CERCA ELETRÔNICA", "INTEGRADOR", "GRUPO", "VISUALIZADO POR",
    "NIVEL", "RÓTULO"
]


def formatar_excel(path_arquivo: str):
    """Formata o arquivo Excel com fonte em negrito e ajusta largura das colunas"""
    try:
        wb = load_workbook(path_arquivo)
        ws = wb.active
        font_bold = Font(bold=True, name="Calibri", size=11)

        # Aplicar fonte em negrito no cabeçalho
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = font_bold

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

        wb.save(path_arquivo)
    except Exception as e:
        print(f"⚠️ Erro ao formatar Excel {path_arquivo}: {e}")


def filtrar_dataframe_por_prefixo(df: pd.DataFrame, prefixos: set, pa: str) -> pd.DataFrame:
    """Filtra o DataFrame pelos prefixos especificados e remove colunas desnecessárias"""
    df_filtrado = df[df["PREFIXO"].isin(prefixos)].copy()
    
    # Remover colunas desnecessárias
    cols_para_remover = [col for col in COLUNAS_REMOVER if col in df_filtrado.columns]
    if cols_para_remover:
        df_filtrado.drop(columns=cols_para_remover, inplace=True)
    
    # Adicionar coluna PA no início
    df_filtrado.insert(0, "PA", pa)
    return df_filtrado


def gerar_graficos(arquivos_filtrados: list):
    """Gera gráficos de análise dos dados com design profissional"""
    try:
        # Concatenar todos os DataFrames
        df_list = []
        for path in arquivos_filtrados:
            try:
                df_temp = pd.read_excel(path)
                df_list.append(df_temp)
            except Exception as e:
                print(f"⚠️ Erro ao ler {os.path.basename(path)}: {e}")
                continue
        
        if not df_list:
            print("❌ Nenhum arquivo válido encontrado para gerar gráficos")
            return None
            
        df_geral = pd.concat(df_list, ignore_index=True)
        print(f"📊 Dados carregados para gráficos: {len(df_geral)} registros")
        
        if df_geral.empty:
            print("❌ Nenhum dado válido encontrado")
            return None
        
        # Configurar estilo do matplotlib
        plt.style.use('default')
        plt.rcParams['font.family'] = 'DejaVu Sans'  # Fonte que suporta mais caracteres
        plt.rcParams['font.size'] = 11
        plt.rcParams['axes.titlesize'] = 14
        plt.rcParams['axes.labelsize'] = 12
        
        # Configurar figura com subplots - tamanho maior e mais espaçamento
        fig = plt.figure(figsize=(18, 14))
        
        # Criar layout personalizado
        gs = fig.add_gridspec(3, 1, height_ratios=[0.8, 2, 2], hspace=0.4)
        
        # Header com informações
        ax_header = fig.add_subplot(gs[0])
        ax_header.axis('off')
        
        # Data atual formatada
        data_atual = datetime.now()
        data_formatada = data_atual.strftime("%d de %B de %Y")
        hora_formatada = data_atual.strftime("%H:%M")
        
        # Meses em português
        meses = {
            'January': 'Janeiro', 'February': 'Fevereiro', 'March': 'Março',
            'April': 'Abril', 'May': 'Maio', 'June': 'Junho',
            'July': 'Julho', 'August': 'Agosto', 'September': 'Setembro',
            'October': 'Outubro', 'November': 'Novembro', 'December': 'Dezembro'
        }
        
        for eng, pt in meses.items():
            data_formatada = data_formatada.replace(eng, pt)
        
        # Header com fundo verde claro sutil
        header_rect = Rectangle((0, 0), 1, 1, transform=ax_header.transAxes, 
                               facecolor='#f0f8f0', edgecolor='#c8e6c9', linewidth=1.5)
        ax_header.add_patch(header_rect)
        
        # Título principal - TEXTO PRETO
        ax_header.text(0.5, 0.7, 'RELATÓRIO EXECUTIVO DE ANÁLISE DE ALERTAS', 
                      ha='center', va='center', fontsize=20, fontweight='bold', 
                      color='#000000', transform=ax_header.transAxes)
        
        # Subtítulo com data - TEXTO PRETO
        ax_header.text(0.5, 0.3, f'Gerado em {data_formatada} às {hora_formatada}', 
                      ha='center', va='center', fontsize=12, 
                      color='#333333', transform=ax_header.transAxes)
        
        # Cores profissionais - tons mais suaves
        cores_tipo = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f']
        cores_pa = ['#66bb6a', '#81c784', '#a5d6a7', '#c8e6c9']  # Verde mais suave

        # Gráfico 1: Contagem de ocorrências por TIPO por PA
        ax1 = fig.add_subplot(gs[1])
        if "TIPO" in df_geral.columns:
            try:
                # Limpar dados de TIPO
                df_geral["TIPO"] = df_geral["TIPO"].astype(str).str.strip()
                df_geral = df_geral[df_geral["TIPO"] != "nan"]
                
                if not df_geral.empty:
                    tipo_contagem = df_geral.groupby(["PA", "TIPO"]).size().unstack(fill_value=0)
                    if not tipo_contagem.empty:
                        # Criar gráfico de barras agrupadas
                        bars = tipo_contagem.plot(kind="bar", ax=ax1, 
                                                color=cores_tipo[:len(tipo_contagem.columns)],
                                                edgecolor='white', linewidth=1.5, width=0.8)
                        
                        # TÍTULOS E LABELS EM PRETO
                        ax1.set_title('DISTRIBUIÇÃO DE ALERTAS POR CATEGORIA E POSTO', 
                                    fontsize=16, fontweight='bold', pad=25, color='#000000')
                        ax1.set_ylabel('Quantidade de Alertas', fontsize=13, fontweight='bold', color='#000000')
                        ax1.set_xlabel('Posto de Atendimento (PA)', fontsize=13, fontweight='bold', color='#000000')
                        
                        # Grid profissional
                        ax1.grid(axis="y", linestyle="-", alpha=0.2, color='#cccccc')
                        ax1.set_axisbelow(True)
                        
                        # Formatação dos eixos - TEXTO PRETO
                        ax1.tick_params(axis='x', rotation=0, labelsize=11, colors='#000000')
                        ax1.tick_params(axis='y', labelsize=10, colors='#000000')
                        
                        # Legenda profissional - TEXTO PRETO
                        legend = ax1.legend(bbox_to_anchor=(1.02, 1), loc='upper left', 
                                          fontsize=10, frameon=True, fancybox=True, 
                                          shadow=True, title="Categorias de Alerta",
                                          title_fontsize=11)
                        legend.get_title().set_fontweight('bold')
                        legend.get_title().set_color('#000000')
                        
                        # Adicionar valores nas barras
                        for container in ax1.containers:
                            ax1.bar_label(container, fmt='%d', fontsize=9, 
                                        fontweight='bold', color='white')
                        
                        # Fundo do gráfico
                        ax1.set_facecolor('#fafafa')
                        
                    else:
                        ax1.text(0.5, 0.5, 'Sem dados de categoria válidos', 
                               ha='center', va='center', transform=ax1.transAxes,
                               fontsize=14, color='#d32f2f')
                        ax1.set_title("DISTRIBUIÇÃO DE ALERTAS - SEM DADOS", color='#000000')
                else:
                    ax1.text(0.5, 0.5, 'Sem dados válidos após processamento', 
                           ha='center', va='center', transform=ax1.transAxes,
                           fontsize=14, color='#d32f2f')
                    ax1.set_title("DISTRIBUIÇÃO DE ALERTAS - SEM DADOS", color='#000000')
            except Exception as e:
                print(f"⚠️ Erro no gráfico de tipos: {e}")
                ax1.text(0.5, 0.5, f'Erro no processamento: {str(e)}', 
                       ha='center', va='center', transform=ax1.transAxes,
                       fontsize=12, color='#d32f2f')
                ax1.set_title("DISTRIBUIÇÃO DE ALERTAS - ERRO", color='#000000')

        # Gráfico 2: Total por PA com destaque especial
        ax2 = fig.add_subplot(gs[2])
        try:
            pa_contagem = df_geral.groupby("PA").size().sort_values(ascending=False)
            if not pa_contagem.empty:
                # Criar gráfico com cores verdes suaves
                bars = ax2.bar(pa_contagem.index, pa_contagem.values, 
                             color=cores_pa[:len(pa_contagem)], 
                             edgecolor='white', linewidth=2.5, width=0.7,
                             alpha=0.8)
                
                # Adicionar efeito de destaque nas barras
                for i, bar in enumerate(bars):
                    height = bar.get_height()
                    # Destacar a barra com maior valor com verde mais intenso
                    if i == 0:  # Primeira barra (maior valor)
                        bar.set_facecolor('#4caf50')  # Verde mais intenso apenas para destaque
                        bar.set_edgecolor('#2e7d32')
                        bar.set_linewidth(3)
                        bar.set_alpha(0.9)
                
                # TÍTULOS E LABELS EM PRETO
                ax2.set_title('VOLUME TOTAL DE ALERTAS POR POSTO DE ATENDIMENTO', 
                            fontsize=16, fontweight='bold', pad=25, color='#000000')
                ax2.set_ylabel('Total de Alertas', fontsize=13, fontweight='bold', color='#000000')
                ax2.set_xlabel('Posto de Atendimento (PA)', fontsize=13, fontweight='bold', color='#000000')
                
                # Grid profissional
                ax2.grid(axis="y", linestyle="-", alpha=0.2, color='#cccccc')
                ax2.set_axisbelow(True)
                
                # Formatação dos eixos - TEXTO PRETO
                ax2.tick_params(axis='x', rotation=0, labelsize=12, colors='#000000')
                ax2.tick_params(axis='y', labelsize=11, colors='#000000')
                
                # Adicionar valores nas barras - TEXTO PRETO
                for i, (pa, valor) in enumerate(pa_contagem.items()):
                    color = '#000000'  # Todos os valores em preto
                    fontsize = 13 if i == 0 else 11
                    ax2.text(i, valor + max(pa_contagem.values) * 0.02, 
                           f'{valor:,}', ha='center', va='bottom', 
                           fontsize=fontsize, fontweight='bold', color=color)
                
                # Fundo do gráfico
                ax2.set_facecolor('#fafafa')
                
                # Linha de média com cor laranja suave
                media = pa_contagem.mean()
                ax2.axhline(y=media, color='#ff9800', linestyle='--', 
                          linewidth=2, alpha=0.7, label=f'Média: {media:.0f} alertas')
                
                # Legenda da média
                legend = ax2.legend(loc='upper right', fontsize=11, frameon=True, 
                                  fancybox=True, shadow=True)
                legend.get_frame().set_facecolor('#ffffff')
                legend.get_frame().set_alpha(0.9)
                
            else:
                ax2.text(0.5, 0.5, 'Sem dados válidos para análise', 
                       ha='center', va='center', transform=ax2.transAxes,
                       fontsize=14, color='#d32f2f')
                ax2.set_title("VOLUME TOTAL DE ALERTAS - SEM DADOS", color='#000000')
        except Exception as e:
            print(f"⚠️ Erro no gráfico de contagem: {e}")
            ax2.text(0.5, 0.5, f'Erro no processamento: {str(e)}', 
                   ha='center', va='center', transform=ax2.transAxes,
                   fontsize=12, color='#d32f2f')
            ax2.set_title("VOLUME TOTAL DE ALERTAS - ERRO", color='#000000')

        # Footer profissional com fundo verde muito claro
        footer_y = 0.02
        footer_rect = Rectangle((0.05, footer_y), 0.9, 0.06, transform=fig.transFigure, 
                               facecolor='#f0f8f0', edgecolor='#c8e6c9', linewidth=1)
        fig.patches.append(footer_rect)
        
        # Informações do footer
        total_registros = len(df_geral)
        total_pas = df_geral['PA'].nunique()
        
        footer_text = (f'Sistema de Monitoramento de Alertas | '
                      f'Total de Registros: {total_registros:,} | '
                      f'Postos Analisados: {total_pas} | '
                      f'Período de Análise: {data_formatada}')
        
        # FOOTER EM PRETO
        fig.text(0.5, footer_y + 0.03, footer_text, 
                ha='center', va='center', fontsize=10, 
                color='#000000', fontweight='bold')
        
        # Ajustar layout
        plt.subplots_adjust(left=0.08, right=0.92, top=0.92, bottom=0.12)
        
        # Salvar na pasta Downloads
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        caminho_saida = os.path.join(downloads_path, "grafico_alertas.png")
        
        # Salvar com alta qualidade
        plt.savefig(caminho_saida, dpi=300, bbox_inches='tight', 
                   facecolor='white', edgecolor='none', format='png',
                   pad_inches=0.2)
        plt.close()  # Importante: fechar a figura para liberar memória
        
        print(f"✅ Gráficos salvos em: {caminho_saida}")
        return caminho_saida
        
    except Exception as e:
        print(f"❌ Erro geral ao gerar gráficos: {e}")
        return None


def processar_arquivo(filepath: str, progress_callback=None) -> list:
    """
    Processa o arquivo CSV principal e gera arquivos Excel filtrados por PA
    
    Args:
        filepath: Caminho para o arquivo CSV
        progress_callback: Função de callback para atualizar progresso
        
    Returns:
        Lista de caminhos dos arquivos gerados
    """
    try:
        print(f"📂 Processando arquivo: {os.path.basename(filepath)}")
        
        # Ler arquivo CSV
        df = pd.read_csv(filepath, encoding="latin1", sep=None, engine="python")
        print(f"📊 Dados carregados: {len(df)} registros")
        
        # Criar pasta de exportação
        pasta_exportados = os.path.join(os.path.dirname(filepath), "exportados")
        os.makedirs(pasta_exportados, exist_ok=True)

        arquivos_gerados = []
        total_grupos = len(PREFIXOS)

        # Processar cada PA
        for i, (pa, prefixos) in enumerate(PREFIXOS.items(), start=1):
            try:
                print(f"🔄 Processando {pa}... ({i}/{total_grupos})")
                
                # Filtrar dados
                df_filtrado = filtrar_dataframe_por_prefixo(df, prefixos, pa)
                
                if df_filtrado.empty:
                    print(f"⚠️ Nenhum dado encontrado para {pa}")
                    continue
                
                # Gerar nome do arquivo com timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:21]
                nome_arquivo = f"veiculos_{pa.lower()}_filtrado_{timestamp}.xlsx"
                caminho_saida = os.path.join(pasta_exportados, nome_arquivo)

                # Salvar Excel
                df_filtrado.to_excel(caminho_saida, index=False)
                formatar_excel(caminho_saida)
                arquivos_gerados.append(caminho_saida)
                
                print(f"✅ {pa}: {len(df_filtrado)} registros salvos")

                # Atualizar progresso
                if progress_callback:
                    try:
                        progress_callback(i / total_grupos)
                    except Exception as callback_error:
                        print(f"⚠️ Erro no callback de progresso: {callback_error}")

                time.sleep(0.1)  # Pequena pausa para não sobrecarregar
                
            except Exception as pa_error:
                print(f"❌ Erro ao processar {pa}: {pa_error}")
                continue

        # Gerar gráficos se houver arquivos
        if arquivos_gerados:
            print("📈 Gerando gráficos...")
            gerar_graficos(arquivos_gerados)
        else:
            print("⚠️ Nenhum arquivo foi gerado, não é possível criar gráficos")

        print(f"🎉 Processamento concluído! {len(arquivos_gerados)} arquivos gerados")
        return arquivos_gerados
        
    except Exception as e:
        print(f"❌ Erro geral no processamento: {e}")
        return []


class ProcessadorUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de Dados - Versão Corrigida")
        self.root.geometry("600x400")
        
        self.arquivo_selecionado = None
        self.processando = False
        
        self.criar_interface()
    
    def criar_interface(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Seleção de arquivo
        ttk.Label(main_frame, text="Selecione o arquivo CSV:").grid(row=0, column=0, sticky=tk.W, pady=5)
        
        frame_arquivo = ttk.Frame(main_frame)
        frame_arquivo.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.label_arquivo = ttk.Label(frame_arquivo, text="Nenhum arquivo selecionado", 
                                      foreground="gray")
        self.label_arquivo.grid(row=0, column=0, sticky=tk.W)
        
        ttk.Button(frame_arquivo, text="Selecionar Arquivo", 
                  command=self.selecionar_arquivo).grid(row=0, column=1, padx=(10, 0))
        
        # Barra de progresso
        ttk.Label(main_frame, text="Progresso:").grid(row=2, column=0, sticky=tk.W, pady=(20, 5))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(main_frame, variable=self.progress_var, 
                                           maximum=100, length=400)
        self.progress_bar.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.label_status = ttk.Label(main_frame, text="Pronto para processar")
        self.label_status.grid(row=4, column=0, columnspan=2, pady=5)
        
        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.grid(row=5, column=0, columnspan=2, pady=20)
        
        self.btn_processar = ttk.Button(frame_botoes, text="Processar Arquivo", 
                                       command=self.iniciar_processamento)
        self.btn_processar.grid(row=0, column=0, padx=5)
        
        ttk.Button(frame_botoes, text="Limpar", 
                  command=self.limpar).grid(row=0, column=1, padx=5)
        
        # Área de log
        ttk.Label(main_frame, text="Log de atividades:").grid(row=6, column=0, sticky=tk.W, pady=(20, 5))
        
        frame_log = ttk.Frame(main_frame)
        frame_log.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        self.text_log = tk.Text(frame_log, height=10, width=70)
        scrollbar = ttk.Scrollbar(frame_log, orient=tk.VERTICAL, command=self.text_log.yview)
        self.text_log.configure(yscrollcommand=scrollbar.set)
        
        self.text_log.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Configurar redimensionamento
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(7, weight=1)
        frame_log.columnconfigure(0, weight=1)
        frame_log.rowconfigure(0, weight=1)
    
    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo CSV",
            filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_selecionado = arquivo
            self.label_arquivo.config(text=os.path.basename(arquivo), foreground="black")
            self.log(f"Arquivo selecionado: {os.path.basename(arquivo)}")
    
    def log(self, mensagem):
        """Adiciona mensagem ao log de forma thread-safe"""
        def _log():
            self.text_log.insert(tk.END, f"{mensagem}\n")
            self.text_log.see(tk.END)
        
        if threading.current_thread() == threading.main_thread():
            _log()
        else:
            self.root.after(0, _log)
    
    def atualizar_progresso(self, valor):
        """Atualiza a barra de progresso de forma thread-safe"""
        def _atualizar():
            self.progress_var.set(valor * 100)
            self.label_status.config(text=f"Processando... {valor*100:.1f}%")
        
        if threading.current_thread() == threading.main_thread():
            _atualizar()
        else:
            self.root.after(0, _atualizar)
    
    def mostrar_erro(self, mensagem):
        """Mostra erro de forma thread-safe"""
        def _mostrar():
            messagebox.showerror("Erro", mensagem)
            self.finalizar_processamento()
        
        if threading.current_thread() == threading.main_thread():
            _mostrar()
        else:
            self.root.after(0, _mostrar)
    
    def mostrar_sucesso(self, mensagem):
        """Mostra sucesso de forma thread-safe"""
        def _mostrar():
            messagebox.showinfo("Sucesso", mensagem)
            self.finalizar_processamento()
        
        if threading.current_thread() == threading.main_thread():
            _mostrar()
        else:
            self.root.after(0, _mostrar)
    
    def iniciar_processamento(self):
        if not self.arquivo_selecionado:
            messagebox.showwarning("Aviso", "Selecione um arquivo CSV primeiro!")
            return
        
        if self.processando:
            messagebox.showwarning("Aviso", "Já existe um processamento em andamento!")
            return
        
        self.processando = True
        self.btn_processar.config(state="disabled")
        self.progress_var.set(0)
        self.label_status.config(text="Iniciando processamento...")
        self.log("=== INICIANDO PROCESSAMENTO ===")
        
        # Executar processamento em thread separada
        thread = threading.Thread(target=self.processar_arquivo_thread)
        thread.daemon = True
        thread.start()
    
    def processar_arquivo_thread(self):
        """Executa o processamento em thread separada"""
        try:
            # Redirecionar prints para o log da interface
            import sys
            from io import StringIO
            
            # Capturar saída do console
            old_stdout = sys.stdout
            sys.stdout = captured_output = StringIO()
            
            try:
                arquivos_gerados = processar_arquivo(
                    self.arquivo_selecionado, 
                    progress_callback=self.atualizar_progresso
                )
                
                # Restaurar stdout e capturar mensagens
                sys.stdout = old_stdout
                output_messages = captured_output.getvalue()
                
                # Adicionar mensagens ao log
                for linha in output_messages.strip().split('\n'):
                    if linha.strip():
                        self.log(linha)
                
                if arquivos_gerados:
                    mensagem = f"Processamento concluído com sucesso!\n{len(arquivos_gerados)} arquivos gerados."
                    self.log(f"✅ {mensagem}")
                    self.mostrar_sucesso(mensagem)
                else:
                    mensagem = "Nenhum arquivo foi gerado. Verifique os dados de entrada."
                    self.log(f"⚠️ {mensagem}")
                    self.mostrar_erro(mensagem)
                    
            finally:
                # Garantir que stdout seja restaurado
                sys.stdout = old_stdout
                
        except Exception as e:
            # CORREÇÃO: Usar a variável 'e' corretamente no lambda
            mensagem_erro = f"Erro durante o processamento: {str(e)}"
            self.log(f"❌ {mensagem_erro}")
            # Correção aplicada: capturar 'e' no escopo do lambda
            self.root.after(0, lambda erro=e: self.mostrar_erro(str(erro)))
    
    def finalizar_processamento(self):
        """Finaliza o processamento e reabilita a interface"""
        self.processando = False
        self.btn_processar.config(state="normal")
        self.progress_var.set(0)
        self.label_status.config(text="Pronto para processar")
        self.log("=== PROCESSAMENTO FINALIZADO ===\n")
    
    def limpar(self):
        if self.processando:
            messagebox.showwarning("Aviso", "Não é possível limpar durante o processamento!")
            return
        
        self.arquivo_selecionado = None
        self.label_arquivo.config(text="Nenhum arquivo selecionado", foreground="gray")
        self.progress_var.set(0)
        self.label_status.config(text="Pronto para processar")
        self.text_log.delete(1.0, tk.END)


def main():
    root = tk.Tk()
    app = ProcessadorUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
