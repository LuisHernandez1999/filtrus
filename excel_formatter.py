"""
Módulo para formatação de arquivos Excel
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook


class ExcelFormatter:
    """Classe responsável pela formatação de arquivos Excel"""
    
    def __init__(self):
        self.font_bold = Font(bold=True, name="Calibri", size=11)
    
    def formatar_arquivo(self, path_arquivo: str) -> bool:
        """
        Formata o arquivo Excel com fonte em negrito e ajusta largura das colunas
        
        Args:
            path_arquivo: Caminho para o arquivo Excel
            
        Returns:
            bool: True se formatado com sucesso, False caso contrário
        """
        try:
            wb = load_workbook(path_arquivo)
            ws = wb.active
            
            # Aplicar fonte em negrito no cabeçalho
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = self.font_bold

            # Ajustar largura das colunas
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

            wb.save(path_arquivo)
            return True
            
        except Exception as e:
            print(f"⚠️ Erro ao formatar Excel {path_arquivo}: {e}")
            return False
